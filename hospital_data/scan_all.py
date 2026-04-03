"""
Scan all Excel files in a folder, identify the summary sheet of each,
and generate a complete FILE_CONFIG block ready to paste into
nhs_data_cleaning_v2.py.

Supports all NHS HES filename formats:
  98-99, 99-00, 2012-13, 2019-20-tab_supp, etc.
"""

import openpyxl
import os
import re

# ============================================================
# ★ Set the folder path containing your Excel files.
#   Leave as "./" if this script is in the same folder.
# ============================================================
DATA_FOLDER = r"./"


def extract_year(filename):
    """
    Extract a standardised year label from a filename.
    Supports both two-digit (98-99) and four-digit (2012-13) formats.
    Returns a string such as '1998-99' or '2012-13', or None if not found.
    """
    fname = filename.lower()
    # Four-digit year: 2012-13, 2019-20
    m = re.search(r'(20\d{2})[\-_](0?\d|1[0-4])', fname)
    if m:
        y1 = m.group(1)
        y2 = str(int(y1) + 1)[-2:]
        return f"{y1}-{y2}"
    # Two-digit year: 98-99, 00-01, 11-12
    m2 = re.search(r'\b(\d{2})[\-_](\d{2})\b', fname)
    if m2:
        y1 = int(m2.group(1))
        full_y1 = (1900 + y1) if y1 >= 98 else (2000 + y1)
        return f"{full_y1}-{str(full_y1 + 1)[-2:]}"
    return None


def is_summary_file(filename):
    """
    Return True if the file is a summary-level file.
    Excludes 3-character (3cha) and 4-character (4cha) diagnosis files,
    which contain more granular data not needed for this visualisation.
    """
    fname = filename.lower()
    if re.search(r'[\-_](3|4)(c|ch|cha|char)([\-_\.]|$)', fname):
        return False
    if re.search(r'diag[\-_](3|4)cha', fname):
        return False
    return True


def find_summary_sheet(sheetnames):
    """
    Select the most likely summary sheet from a workbook's sheet list.
    Prefers sheets containing 'summary'; falls back to sheets containing
    'diag' (excluding introduction sheets).
    """
    for s in sheetnames:
        if 'summary' in s.lower():
            return s
    for s in sheetnames:
        if 'diag' in s.lower() and 'intro' not in s.lower():
            return s
    return sheetnames[0]


# ============================================================
# Main scan
# ============================================================
print("=" * 70)
print("Scanning all Excel files -> generating FILE_CONFIG")
print("=" * 70)

all_files = sorted([f for f in os.listdir(DATA_FOLDER)
                    if f.lower().endswith('.xlsx') or f.lower().endswith('.xls')])
print(f"Found {len(all_files)} Excel file(s) in folder\n")

results, skipped, no_year = [], [], []

for fname in all_files:
    # Skip 3cha / 4cha files
    if not is_summary_file(fname):
        skipped.append(fname)
        continue

    year = extract_year(fname)
    if year is None:
        no_year.append(fname)
        print(f"  WARNING  Cannot detect year: {fname}")
        continue

    fpath = os.path.join(DATA_FOLDER, fname)
    try:
        wb    = openpyxl.load_workbook(fpath, read_only=True)
        sheet = find_summary_sheet(wb.sheetnames)
        results.append((year, fname, sheet))
        print(f"  OK  {year}  |  sheet: '{sheet}'")
    except Exception as e:
        no_year.append(fname)
        print(f"  ERROR  {fname}: {e}")

# Sort by year label
results.sort(key=lambda x: x[0])
years      = [r[0] for r in results]
duplicates = {y for y in years if years.count(y) > 1}

# ============================================================
# Print FILE_CONFIG block
# ============================================================
print("\n" + "=" * 70)
print("Copy the block below and replace FILE_CONFIG in nhs_data_cleaning_v2.py")
print("=" * 70 + "\n")

print("FILE_CONFIG = {")
for year, fname, sheet in results:
    note = "  # WARNING: duplicate year - keep only one line" if year in duplicates else ""
    print(f'    "{year}": ("{fname}", "{sheet}"),{note}')
print("}")

# ============================================================
# Summary
# ============================================================
print(f"\n{'='*70}")
print(f"Successfully identified : {len(results)} year(s)")
print(f"Skipped (3cha / 4cha)  : {len(skipped)} file(s)")

if no_year:
    print(f"\nWARNING  The following file(s) could not be processed automatically:")
    for f in no_year:
        print(f"   {f}")

if duplicates:
    print(f"\nWARNING  Duplicate year(s) detected - manually remove one line per year:")
    for y in sorted(duplicates):
        for yr, f, s in results:
            if yr == y:
                print(f"   {y}: {f}  (sheet: {s})")
